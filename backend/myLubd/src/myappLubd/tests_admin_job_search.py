from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.test import RequestFactory, TestCase

from .admin import IsDefectFilter, JobAdmin, _excel_image_for_export
from .models import Job, Property, Room


User = get_user_model()


class JobAdminSearchTests(TestCase):
    def setUp(self):
        self.request = RequestFactory().get('/admin/myappLubd/job/')
        self.admin = JobAdmin(Job, AdminSite())
        self.user = User.objects.create_user(username='engineer', password='pw12345!')
        self.property = Property.objects.create(name='Admin Search Hotel')
        self.room = Room.objects.create(
            room_id=987654,
            name='LUBD-1205',
            room_type='Deluxe',
            property=self.property,
        )
        self.job = Job.objects.create(
            user=self.user,
            property=self.property,
            description='Replace air filter',
            remarks='Admin search test',
            status='pending',
            priority='medium',
        )
        self.job.job_id = 'jADMINSEARCH'
        self.job.save(update_fields=['job_id'])
        self.job.rooms.add(self.room)

    def test_search_matches_job_id(self):
        queryset, _ = self.admin.get_search_results(
            self.request,
            Job.objects.all(),
            'ADMINSEARCH',
        )

        self.assertIn(self.job, queryset)

    def test_search_matches_room_name(self):
        queryset, _ = self.admin.get_search_results(self.request, Job.objects.all(), '1205')

        self.assertIn(self.job, queryset)

    def test_search_does_not_match_room_id(self):
        queryset, _ = self.admin.get_search_results(
            self.request,
            Job.objects.all(),
            str(self.room.room_id),
        )

        self.assertNotIn(self.job, queryset)

    def test_search_matches_description(self):
        queryset, _ = self.admin.get_search_results(
            self.request,
            Job.objects.all(),
            'Replace air filter',
        )

        # Description is an intentional JobAdmin search field.
        self.assertIn(self.job, queryset)


class IsDefectFilterTests(TestCase):
    def setUp(self):
        self.request = RequestFactory().get('/admin/myappLubd/job/?is_defect=1')
        self.user = User.objects.create_user(username='engineer-filter', password='pw12345!')
        self.property = Property.objects.create(name='Admin Filter Hotel')
        self.defect_job = Job.objects.create(
            user=self.user,
            property=self.property,
            description='Defect job',
            remarks='Admin filter test',
            status='pending',
            priority='medium',
            is_defective=True,
        )
        self.non_defect_job = Job.objects.create(
            user=self.user,
            property=self.property,
            description='Non defect job',
            remarks='Admin filter test',
            status='pending',
            priority='medium',
            is_defective=False,
        )

    def test_is_defect_filter_uses_requested_query_parameter(self):
        defect_filter = IsDefectFilter(
            self.request,
            {'is_defect': '1'},
            Job,
            JobAdmin(Job, AdminSite()),
        )

        queryset = defect_filter.queryset(self.request, Job.objects.all())

        self.assertIn(self.defect_job, queryset)
        self.assertNotIn(self.non_defect_job, queryset)


class JobAdminCsvExportTests(TestCase):
    def setUp(self):
        self.request = RequestFactory().get('/admin/myappLubd/job/')
        self.admin = JobAdmin(Job, AdminSite())
        self.user = User.objects.create_user(username='csv-engineer', password='pw12345!')
        self.property = Property.objects.create(name='Admin CSV Export Hotel')
        self.job = Job.objects.create(
            user=self.user,
            property=self.property,
            description='CSV export image test',
            remarks='Includes image URL',
            status='pending',
            priority='medium',
        )

    @staticmethod
    def image_upload(name, *, format='JPEG'):
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        from PIL import Image

        content = BytesIO()
        Image.new('RGB', (2, 2), color='red').save(content, format=format)
        return SimpleUploadedFile(name, content.getvalue(), content_type='image/jpeg')

    def test_export_jobs_csv_includes_image_urls_and_display_formulas(self):
        from csv import DictReader
        from io import StringIO
        image = self.job.job_images.create(
            uploaded_by=self.user,
            image=self.image_upload('before.jpg'),
        )

        response = self.admin.export_jobs_csv(self.request, Job.objects.filter(pk=self.job.pk))
        rows = list(DictReader(StringIO(response.content.decode())))

        expected_url = self.request.build_absolute_uri(image.image.url)
        self.assertEqual(rows[0]['Image URLs'], expected_url)
        self.assertEqual(
            rows[0]['Image Formulas (Excel/Google Sheets)'],
            f'=IMAGE("{expected_url}")',
        )
        self.assertIn('CSV cannot embed images', rows[0]['Image Export Notes'])

    def test_export_jobs_google_sheets_csv_uses_image_formulas(self):
        from csv import DictReader
        from io import StringIO
        image = self.job.job_images.create(
            uploaded_by=self.user,
            image=self.image_upload('before-for-sheets.jpg'),
        )

        response = self.admin.export_jobs_google_sheets_csv(self.request, Job.objects.filter(pk=self.job.pk))
        rows = list(DictReader(StringIO(response.content.decode())))

        expected_url = self.request.build_absolute_uri(image.image.url)
        self.assertIn('jobs_google_sheets_', response['Content-Disposition'])
        self.assertEqual(rows[0]['Image URLs'], expected_url)
        self.assertEqual(
            rows[0]['Image Formulas (Excel/Google Sheets)'],
            f'=IMAGE("{expected_url}")',
        )

    def test_export_jobs_excel_leaves_mpo_images_as_urls(self):
        from unittest.mock import patch

        image = self.job.job_images.create(
            uploaded_by=self.user,
            image=self.image_upload('stereo.mpo'),
        )

        with patch('openpyxl.drawing.image.Image', side_effect=KeyError('.mpo')):
            response = self.admin.export_jobs_excel(self.request, Job.objects.filter(pk=self.job.pk))

        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        from io import BytesIO
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(response.content))
        row = next(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(row[16], 'Image URL only (unsupported Excel preview)')
        self.assertEqual(row[17], self.request.build_absolute_uri(image.image.url))

    def test_excel_image_conversion_opens_mpo_to_classify_it(self):
        from pathlib import Path
        from tempfile import TemporaryDirectory
        from unittest.mock import patch

        class CapturingDrawingImage:
            def __init__(self, image_buffer):
                self.image_buffer = image_buffer

        with TemporaryDirectory() as temporary_directory:
            image_path = Path(temporary_directory) / 'stereo.mpo'
            image_path.write_bytes(b'not-real-mpo-image-bytes')

            with patch('PIL.Image.open') as image_open:
                with self.assertRaises(ValueError):
                    _excel_image_for_export(str(image_path), CapturingDrawingImage)

            image_open.assert_called_once_with(str(image_path))

    def test_excel_image_conversion_uses_thumbnail_preview(self):
        from io import BytesIO
        from pathlib import Path
        from tempfile import TemporaryDirectory
        from PIL import Image as PILImage

        class CapturingDrawingImage:
            def __init__(self, image_buffer):
                self.image_buffer = image_buffer

        with TemporaryDirectory() as temporary_directory:
            image_path = Path(temporary_directory) / 'large-preview.bmp'
            PILImage.new('RGB', (800, 600), color='blue').save(image_path, format='BMP')

            excel_image, converted_buffer = _excel_image_for_export(
                str(image_path),
                CapturingDrawingImage,
            )

            self.assertIs(excel_image.image_buffer, converted_buffer)
            converted_buffer.seek(0)
            with PILImage.open(BytesIO(converted_buffer.getvalue())) as converted_image:
                self.assertEqual(converted_image.format, 'PNG')
                self.assertLessEqual(converted_image.width, 120)
                self.assertLessEqual(converted_image.height, 90)


    def test_excel_image_conversion_rejects_invalid_image_data(self):
        from pathlib import Path
        from tempfile import TemporaryDirectory
        from unittest.mock import patch

        class CapturingDrawingImage:
            def __init__(self, image_buffer):
                self.image_buffer = image_buffer

        with TemporaryDirectory() as temporary_directory:
            image_path = Path(temporary_directory) / 'oversized.bmp'
            image_path.write_bytes(b'0')

            from PIL import UnidentifiedImageError

            with self.assertRaises(UnidentifiedImageError):
                _excel_image_for_export(str(image_path), CapturingDrawingImage)

    def test_export_jobs_excel_embeds_normalized_image_uploads(self):
        from io import BytesIO
        from openpyxl import load_workbook
        image = self.job.job_images.create(
            uploaded_by=self.user,
            image=self.image_upload('camera-upload.jfif'),
        )

        response = self.admin.export_jobs_excel(self.request, Job.objects.filter(pk=self.job.pk))

        workbook = load_workbook(BytesIO(response.content))
        row = next(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(row[16], 'Embedded')
        self.assertEqual(row[17], self.request.build_absolute_uri(image.image.url))
